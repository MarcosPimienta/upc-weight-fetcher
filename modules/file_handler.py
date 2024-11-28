import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def load_excel(file_path, sheet_name=0, header_row=None):
    """
    Loads data from a specific sheet and header row in an Excel file into a DataFrame.

    Parameters:
    - file_path: Path to the Excel file.
    - sheet_name: The sheet name or index to load data from.
    - header_row: Row number (0-indexed) to use as the header. None if the first row.

    Returns:
    - DataFrame with the loaded data.
    """
    return pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)

def save_to_excel(df, file_path):
    """Saves DataFrame to an Excel file."""
    df.to_excel(file_path, index=False)

def save_to_csv(df, file_path):
    """Saves DataFrame to a CSV file."""
    df.to_csv(file_path, index=False)

def sanitize_dataframe(df):
    """
    Replaces problematic characters in the DataFrame to avoid errors during saving.

    Parameters:
    - df: The DataFrame to sanitize.

    Returns:
    - A sanitized DataFrame with empty values instead of problematic placeholders.
    """
    return df.applymap(lambda x: "" if pd.isnull(x) or x == "N/A" else x)

def save_to_excel_with_highlighting(df, api_name, na_fields):
    """
    Save the dataframe to an Excel file with highlighting for rows missing all fields.
    Files are saved to API-specific folders.

    Parameters:
    - df: The DataFrame to save.
    - api_name: The name of the API used (e.g., "upcitemdb", "redcircle").
    - na_fields: List of columns to check for missing values.
    """
    # Define the folder based on API name
    folder = f"results/{api_name}"
    os.makedirs(folder, exist_ok=True)  # Create the folder if it doesn't exist

    # Define the file name dynamically based on API
    file_name = f"{api_name}_products.xlsx"
    file_path = os.path.join(folder, file_name)

    # Save DataFrame to Excel
    df.to_excel(file_path, index=False, engine="openpyxl")

    # Apply highlighting for rows with all N/A fields
    wb = load_workbook(file_path)
    ws = wb.active
    highlight = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        # Check if all specified fields are missing
        if all(pd.isna(getattr(row, field, None)) for field in na_fields):
            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = highlight

    wb.save(file_path)
    print(f"File saved to: {file_path}")
